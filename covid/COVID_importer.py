import time
from datetime import timedelta, datetime
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
import string

import sys
sys.path.append("..")

from DataAnalysisProcessor import USER_ORIGIN, FILE_ORIGIN

TYPES = ['Confirmed', 'Deaths', 'Recovered']

def base26(decimal):
    char = []
    while decimal > 0:
        decimal, digit = divmod(decimal, 26)
        char.append(string.ascii_uppercase[digit - 1])
    return ''.join(reversed(char))

class COVID_importer:
    def __init__(self, type, dateString):
        self.type = type
        self.dateString = dateString

    def import_data(self):
        df = self._process_data()
        workbook = load_workbook(self._save_path())
        worksheet = workbook.get_sheet_by_name(self.type)
        workbook.remove_sheet(worksheet)
        # workbook.save(self._save_path())

        writer = pd.ExcelWriter(self._save_path(), engine='openpyxl')
        writer.book = workbook
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
        # workbook.create_sheet(self.type)
        df.to_excel(writer, sheet_name=self.type)
        worksheet = writer.sheets[self.type]
        table = Table(displayName=self.type, ref="A1:{co}{ro}".format(co=base26(df.shape[1] + 1), ro=df.shape[0] + 1))# Add a default style with striped rows and banded columns
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        worksheet.add_table(table)
        writer.save()

    def _process_data(self):
        path = self._origin_path()
        df = pd.read_excel(path)
        grouped_df = df.groupby("Country/Region").sum()
        return grouped_df


    def _origin_path(self):
        return USER_ORIGIN + "/Dropbox (Rosalind Advisors)/Data Analytics Shared Folder/COVID-19/" \
                             "JHU-COVID-{date}-{type}.xlsx".format(date=self.dateString, type=self.type)

    def _save_path(self):
        return USER_ORIGIN + "/Dropbox (Rosalind Advisors)/Data Analytics Shared Folder/COVID-19/" \
                             "Rosalind COVID-19 Analysis.xlsx"


target_date = datetime.today() - timedelta(days=1)
# target_date = datetime(2020, 3, 21)
dateString = "{m}-{d}-{y}".format(m=target_date.month, d=target_date.day, y=target_date.year)
print(dateString)

for datatype in TYPES:
    COVID_importer(datatype, dateString).import_data()
